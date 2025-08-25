"""Excel (XLSX) postback handler."""

import os
from typing import List, Dict, Any
import logging
from openpyxl import Workbook, load_workbook
from .base import PostbackHandler

logger = logging.getLogger(__name__)


class XLSXPostbackHandler(PostbackHandler):
    """Handler that writes enriched rows to an Excel file."""
    
    def __init__(self, config: Dict[str, Any]):
        super().__init__(config)
        self.output_path = config.get('output_path', './outputs/postback.xlsx')
        self.sheet_name = config.get('sheet_name', 'Enriched_Data')
        
    def validate_config(self) -> bool:
        """Validate XLSX handler configuration."""
        if not self.output_path:
            logger.error("XLSX handler missing output_path")
            return False
        return True
        
    def post(self, rows: List[Dict[str, Any]]) -> bool:
        """Write enriched rows to Excel file using pandas to avoid Bad CRC-32 errors.
        
        Args:
            rows: List of enriched data dictionaries
            
        Returns:
            True if successful, False otherwise
        """
        if not rows:
            logger.warning("No rows to write to XLSX")
            return True
            
        # Filter out empty dictionaries and validate data
        valid_rows = [row for row in rows if row and isinstance(row, dict) and any(row.values())]
        
        if not valid_rows:
            logger.warning("No valid data rows to write to XLSX (all rows were empty or invalid)")
            return True
            
        # HOTFIX: Use pandas ExcelWriter exclusively to eliminate Bad CRC-32 errors
        try:
            import pandas as pd
            
            # Ensure output directory exists
            os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
            
            # Remove any existing file to prevent corruption
            if os.path.exists(self.output_path):
                try:
                    os.remove(self.output_path)
                    logger.info(f"Removed existing Excel file to prevent Bad CRC-32: {self.output_path}")
                except Exception as e:
                    logger.warning(f"Could not remove existing file: {e}")
            
            # Convert to DataFrame and export directly
            df = pd.DataFrame(valid_rows)
            
            with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=self.sheet_name, index=False)
            
            logger.info(f"Successfully wrote {len(valid_rows)} rows to {self.output_path} using pandas hotfix")
            return True
            
        except Exception as e:
            logger.error(f"Pandas Excel hotfix failed: {e}")
            return False
            
        try:
            # Ensure output directory exists
            os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
            
            # Always create a fresh workbook to avoid corruption issues
            # This prevents Bad CRC-32 errors from corrupted or locked existing files
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            next_row = 1
            
            # If file exists and we want to preserve data, try to read it first
            # But use a more robust approach with fallback
            existing_data = []
            if os.path.exists(self.output_path):
                try:
                    # Try to read existing data before overwriting
                    temp_wb = load_workbook(self.output_path)
                    if self.sheet_name in temp_wb.sheetnames:
                        temp_ws = temp_wb[self.sheet_name]
                        # Read existing data (skip headers)
                        for row in temp_ws.iter_rows(min_row=2, values_only=True):
                            if any(row):  # Skip empty rows
                                existing_data.append(row)
                        logger.info(f"Preserved {len(existing_data)} existing rows from Excel file")
                    temp_wb.close()
                except Exception as load_error:
                    logger.warning(f"Could not load existing Excel file (will create fresh): {load_error}")
                    # Continue with fresh file - this handles the Bad CRC-32 case
                
                # Remove the potentially corrupted file
                try:
                    os.remove(self.output_path)
                    logger.info(f"Removed existing Excel file to prevent corruption issues")
                except Exception as remove_error:
                    logger.warning(f"Could not remove existing file: {remove_error}")
            
            # Write headers
            headers = list(valid_rows[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            next_row = 2
            
            # Write existing data first (if any)
            if existing_data:
                for existing_row in existing_data:
                    for col, value in enumerate(existing_row, 1):
                        ws.cell(row=next_row, column=col, value=value)
                    next_row += 1
                logger.info(f"Restored {len(existing_data)} existing rows to Excel file")
            
            # Write new data rows
            for row_data in valid_rows:
                for col, key in enumerate(headers, 1):
                    value = row_data.get(key)  # Use .get() to handle missing keys
                    # Handle None values and complex objects
                    if value is None:
                        value = ""
                    elif isinstance(value, (dict, list)):
                        value = str(value)
                    ws.cell(row=next_row, column=col, value=value)
                next_row += 1
                
            wb.save(self.output_path)
            logger.info(f"Successfully wrote {len(valid_rows)} rows to {self.output_path}")
            return True
            
        except Exception as e:
            logger.error(f"OpenPyXL Excel export failed: {e}")
            
            # Fallback: Try pandas Excel export as alternative
            try:
                logger.info("Attempting fallback Excel export using pandas...")
                import pandas as pd
                
                # Convert rows to DataFrame
                df = pd.DataFrame(valid_rows)
                
                # Use pandas ExcelWriter with openpyxl engine
                with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=self.sheet_name, index=False)
                    
                logger.info(f"Fallback Excel export successful: wrote {len(valid_rows)} rows to {self.output_path}")
                return True
                
            except Exception as fallback_error:
                logger.error(f"Fallback Excel export also failed: {fallback_error}")
                
                # Final fallback: Save as CSV with .xlsx extension warning
                try:
                    import pandas as pd
                    df = pd.DataFrame(valid_rows)
                    csv_path = self.output_path.replace('.xlsx', '_excel_fallback.csv')
                    df.to_csv(csv_path, index=False)
                    logger.warning(f"Excel export completely failed - saved as CSV instead: {csv_path}")
                    logger.warning("Users should be notified that Excel file was saved as CSV due to export issues")
                    return True  # Return True since data was saved, just in different format
                    
                except Exception as final_error:
                    logger.error(f"All Excel export methods failed: {final_error}")
                    return False